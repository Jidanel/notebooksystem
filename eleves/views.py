from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from .models import Eleve
from .forms import *
from openpyxl import load_workbook
from django.core.files.storage import FileSystemStorage
from classes.models import *
from .utils import envoyer_sms
import json
from django.db.models import Q
from datetime import datetime
from django.db import IntegrityError
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from utilisateurs.decorators import *
from utilisateurs.views import *
from parametres.models import *
from django.core.cache import cache

@role_required(allowed_roles=['Admin_', 'SG'])
@login_required
def liste_eleves(request):
    query = request.GET.get('q')
    page_number = request.GET.get('page', 1)

    # Créer une clé de cache unique basée sur la query et la page
    cache_key = f"liste_eleves_{query}_{page_number}"

    # Tenter de récupérer les résultats depuis le cache
    page_obj = cache.get(cache_key)

    if not page_obj:
        eleves = Eleve.objects.select_related('classe_actuelle').all().order_by('nom')

        if query:
            eleves = eleves.filter(nom__icontains=query)

        paginator = Paginator(eleves, 100)
        page_obj = paginator.get_page(page_number)

        # Mettre la page en cache pendant 15 minutes
        cache.set(cache_key, page_obj, timeout=60*60)

    return render(request, 'eleves/liste_eleves.html', {'page_obj': page_obj, 'query': query})


@role_required(allowed_roles=['Admin_', 'SG'])
@login_required
def ajouter_eleve(request):
    if request.method == 'POST':
        form = EleveForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Élève ajouté avec succès.")
            return redirect('ajouter_eleve')
    else:
        form = EleveForm()
    return render(request, 'eleves/ajouter_eleve.html', {'form': form})

@role_required(allowed_roles=['Admin_', 'SG'])
@login_required
def modifier_eleve(request, eleve_id):
    cache_key = f"eleve_{eleve_id}"
    eleve = cache.get(cache_key)

    if not eleve:
        eleve = get_object_or_404(Eleve, id=eleve_id)
        cache.set(cache_key, eleve, timeout=60*15)

    if request.method == 'POST':
        form = EleveForm(request.POST, request.FILES, instance=eleve)
        if form.is_valid():
            form.save()
            cache.delete(cache_key)
            cache.delete_pattern('liste_eleves_*')
            messages.success(request, "Élève modifié avec succès.")
            return redirect('liste_eleves')
    else:
        form = EleveForm(instance=eleve)
    return render(request, 'eleves/modifier_eleve.html', {'form': form})

@role_required(allowed_roles=['Admin_', 'SG'])
@login_required
def confirmer_suppression_eleve(request, eleve_id):
    return supprimer_objet_securise(
        request, 
        model=Eleve, 
        pk=eleve_id, 
        redirect_url='liste_eleves'
    )


@role_required(allowed_roles=['Admin_', 'SG'])
@login_required
def supprimer_eleves_selectionnes(request):
    if request.method == "POST":
        eleve_ids = request.POST.getlist('eleves')
        eleves = Eleve.objects.filter(id__in=eleve_ids)
        
        if not eleves.exists():
            messages.error(request, "Aucun élève sélectionné.")
        else:
            total = eleves.count()
            eleves.delete()
            cache.delete_pattern('liste_eleves_*')
            messages.success(request, f'{total} élèves ont été supprimés avec succès.')
        return redirect('liste_eleves')



@role_required(allowed_roles=['Admin_', 'SG'])
@login_required
def importer_eleves(request):
    if request.method == 'POST':
        classe_id = request.POST['classe']
        classe = Classe.objects.get(id=classe_id)
        file = request.FILES['file']
        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        filepath = fs.path(filename)

        wb = load_workbook(filepath)
        ws = wb.active

        imported_count = 0
        duplicate_count = 0
        error_count = 0
        duplicates = []
        new_eleves = []
        errors = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                nom = row[0] or ''
                prenom = row[1] or ''
                sexe = row[2] or ''
                date_naissance = row[3]
                lieu_naissance = row[4] or ''
                matricule = row[5]
                statut = row[6] or ''
                contact_parent = row[7] or ''

                if not nom or not matricule:
                    raise ValueError(f"Nom et matricule sont obligatoires (Nom: {nom}, Matricule: {matricule})")

                if isinstance(date_naissance, datetime):
                    date_naissance = date_naissance.strftime('%Y-%m-%d')

                eleve = Eleve.objects.filter(matricule=matricule).first()
                if eleve:
                    duplicates.append({
                        'nom': nom,
                        'prenom': prenom,
                        'sexe': sexe,
                        'date_naissance': date_naissance,
                        'lieu_naissance': lieu_naissance,
                        'matricule': matricule,
                        'statut': statut,
                        'contact_parent': contact_parent,
                        'classe_actuelle': classe.id,
                        'existing_eleve': eleve.id
                    })
                    duplicate_count += 1
                else:
                    new_eleves.append({
                        'nom': nom,
                        'prenom': prenom,
                        'sexe': sexe,
                        'date_naissance': date_naissance,
                        'lieu_naissance': lieu_naissance,
                        'matricule': matricule,
                        'statut': statut,
                        'contact_parent': contact_parent,
                        'classe_actuelle': classe.id
                    })
                    imported_count += 1
            except Exception as e:
                errors.append(f"Erreur sur la ligne {row}: {str(e)}")
                error_count += 1

        request.session['new_eleves'] = new_eleves
        request.session['duplicates'] = duplicates
        request.session['errors'] = errors

        if errors:
            return redirect('resolve_import_errors')
        
        if duplicates:
            return redirect('resolve_duplicates')

        for eleve_data in new_eleves:
            classe = Classe.objects.get(id=eleve_data['classe_actuelle'])
            Eleve.objects.create(
                nom=eleve_data['nom'],
                prenom=eleve_data['prenom'],
                sexe=eleve_data['sexe'],
                date_naissance=eleve_data['date_naissance'],
                lieu_naissance=eleve_data['lieu_naissance'],
                matricule=eleve_data['matricule'],
                statut=eleve_data['statut'],
                contact_parent=eleve_data['contact_parent'],
                classe_actuelle=classe
            )

        cache.delete_pattern('liste_eleves_*')
        messages.success(request, f"Importation réussie: {imported_count} élèves importés, {duplicate_count} doublons détectés, {error_count} erreurs.")
        return redirect('liste_eleves')

    classes = Classe.objects.all()
    return render(request, 'eleves/importer_eleves.html', {'classes': classes})

@role_required(allowed_roles=['Admin_', 'SG'])
@login_required
def resolve_import_errors(request):
    errors = request.session.get('errors', [])
    return render(request, 'eleves/resolve_import_errors.html', {'errors': errors})

@role_required(allowed_roles=['Admin_', 'SG'])
@login_required
def resolve_duplicates(request):
    if request.method == 'POST':
        action = request.POST.getlist('action')
        duplicates = request.session.get('duplicates', [])
        new_eleves = request.session.get('new_eleves', [])
        
        errors = []

        for idx, choice in enumerate(action):
            eleve_data = duplicates[idx]
            existing_eleve = Eleve.objects.get(id=eleve_data['existing_eleve'])
            classe = Classe.objects.get(id=eleve_data['classe_actuelle'])

            date_naissance = eleve_data['date_naissance']
            if isinstance(date_naissance, datetime):
                date_naissance = date_naissance.strftime('%Y-%m-%d')

            try:
                if choice == 'fusionner':
                    existing_eleve.nom = eleve_data['nom']
                    existing_eleve.prenom = eleve_data['prenom']
                    existing_eleve.sexe = eleve_data['sexe']
                    existing_eleve.date_naissance = date_naissance
                    existing_eleve.lieu_naissance = eleve_data['lieu_naissance']
                    existing_eleve.statut = eleve_data['statut']
                    existing_eleve.contact_parent = eleve_data['contact_parent']
                    existing_eleve.classe_actuelle = classe
                    existing_eleve.save()
                elif choice == 'dupliquer':
                    eleve_data.pop('existing_eleve')
                    Eleve.objects.create(
                        nom=eleve_data['nom'],
                        prenom=eleve_data['prenom'],
                        sexe=eleve_data['sexe'],
                        date_naissance=date_naissance,
                        lieu_naissance=eleve_data['lieu_naissance'],
                        matricule=eleve_data['matricule'],
                        statut=eleve_data['statut'],
                        contact_parent=eleve_data['contact_parent'],
                        classe_actuelle=classe
                    )
            except IntegrityError as e:
                errors.append(f"Erreur de duplication pour le matricule {eleve_data['matricule']}: {str(e)}")
            except ValueError as e:
                errors.append(f"Erreur de validation pour le matricule {eleve_data['matricule']}: {str(e)}")

        for eleve_data in new_eleves:
            classe = Classe.objects.get(id=eleve_data['classe_actuelle'])
            try:
                date_naissance = eleve_data['date_naissance']
                if isinstance(date_naissance, datetime):
                    date_naissance = date_naissance.strftime('%Y-%m-%d')

                Eleve.objects.create(
                    nom=eleve_data['nom'],
                    prenom=eleve_data['prenom'],
                    sexe=eleve_data['sexe'],
                    date_naissance=date_naissance,
                    lieu_naissance=eleve_data['lieu_naissance'],
                    matricule=eleve_data['matricule'],
                    statut=eleve_data['statut'],
                    contact_parent=eleve_data['contact_parent'],
                    classe_actuelle=classe
                )
            except IntegrityError as e:
                errors.append(f"Erreur de duplication pour le matricule {eleve_data['matricule']}: {str(e)}")
            except ValueError as e:
                errors.append(f"Erreur de validation pour le matricule {eleve_data['matricule']}: {str(e)}")

        if errors:
            messages.error(request, " ".join(errors))
        else:
            messages.success(request, "Doublons résolus et élèves importés avec succès.")
        
        return redirect('liste_eleves')

    duplicates = request.session.get('duplicates', [])
    return render(request, 'eleves/resolve_duplicates.html', {'duplicates': duplicates})

@role_required(allowed_roles=['Admin_', 'SG'])
@login_required
def contacter_parents(request, eleve_id):
    eleve = get_object_or_404(Eleve, id=eleve_id)
    if request.method == 'POST':
        message = request.POST.get('message')
        if message:
            response = envoyer_sms(eleve.contact_parent, message)
            if response['success']:
                messages.success(request, "Message envoyé avec succès.")
            else:
                messages.error(request, "Échec de l'envoi du message.")
        return redirect('liste_eleves')
    return render(request, 'eleves/contacter_parents.html', {'eleve': eleve})

@role_required(allowed_roles=['Admin_', 'SG'])
@login_required
def liste_eleves_par_classe(request):
    classes = Classe.objects.all()
    selected_classe_id = request.GET.get('classe_id')
    selected_classe = None
    eleves = None
    stats = {}
    
    if selected_classe_id:
        selected_classe = get_object_or_404(Classe, id=selected_classe_id)
        eleves = Eleve.objects.filter(classe_actuelle=selected_classe).order_by('nom')

        paginator = Paginator(eleves, 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        stats['total_eleves'] = eleves.count()
        stats['total_eleves_redoublants'] = eleves.filter(statut='Redoublant').count()
        stats['total_eleves_nouveaux'] = eleves.filter(statut='Nouveau').count()
        stats['total_garcons'] = eleves.filter(sexe='Masculin').count()
        stats['total_filles'] = eleves.filter(sexe='Feminin').count()
        stats['garcons_redoublants'] = eleves.filter(sexe='Masculin', statut='Redoublant').count()
        stats['garcons_nouveaux'] = eleves.filter(sexe='Masculin', statut='Nouveau').count()
        stats['filles_redoublantes'] = eleves.filter(sexe='Feminin', statut='Redoublant').count()
        stats['filles_nouvelles'] = eleves.filter(sexe='Feminin', statut='Nouveau').count()

    context = {
        'classes': classes,
        'selected_classe': selected_classe,
        'page_obj': page_obj if eleves else None,
        'stats': stats
    }
    return render(request, 'eleves/liste_eleves_par_classe.html', context)

@role_required(allowed_roles=['Admin_', 'SG'])
@login_required
def imprimer_liste_eleves(request, classe_id):
    cache_key = f"imprimer_liste_eleves_{classe_id}"
    classe = get_object_or_404(Classe, id=classe_id)
    eleves = cache.get(cache_key)

    if not eleves:
        eleves = Eleve.objects.filter(classe_actuelle=classe).select_related('classe_actuelle').order_by('nom')
        cache.set(cache_key, eleves, timeout=60*15)

    parametres_etablissement = ParametresEtablissement.objects.first()

    stats = {
        'total_eleves': eleves.count(),
        'total_eleves_redoublants': eleves.filter(statut='Redoublant').count(),
        'total_eleves_nouveaux': eleves.filter(statut='Nouveau').count(),
        'total_garcons': eleves.filter(sexe='Masculin').count(),
        'total_filles': eleves.filter(sexe='Feminin').count(),
        'garcons_redoublants': eleves.filter(sexe='Masculin', statut='Redoublant').count(),
        'garcons_nouveaux': eleves.filter(sexe='Masculin', statut='Nouveau').count(),
        'filles_redoublantes': eleves.filter(sexe='Feminin', statut='Redoublant').count(),
        'filles_nouvelles': eleves.filter(sexe='Feminin', statut='Nouveau').count(),
    }

    for eleve in eleves:
        if eleve.sexe == 'Masculin':
            eleve.sexe = 'M'
        elif eleve.sexe == 'Feminin':
            eleve.sexe = 'F'
    
        if eleve.statut == 'Redoublant':
            eleve.statut = 'R'
        elif eleve.statut == 'Nouveau':
            eleve.statut = 'N'

    template_path = 'eleves/imprimer_liste_eleves.html'
    context = {
        'classe': classe,
        'eleves': eleves,
        'stats': stats,
        'parametres_etablissement': parametres_etablissement
    }
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Liste_eleves_{classe.nom}.pdf"'
    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Une erreur s\'est produite lors de la génération du PDF', status=400)
    return response